import openpyxl
import json

wb = openpyxl.load_workbook(filename="data//table_of_trial_data_for_radar_diagram.xlsx", data_only=True)
ws = wb.active

data = {

}

# value = int(ws.max_column)
# print(ws.max_column)

print(ws.max_row ,ws["AP23"].value)
print(ws.max_column, ws["AP27"].value)
MaxRow = data["MaxRow"] = int(ws["AP23"].value)
MaxColumn = data["MaxColumn"] = int(ws["AP27"].value)

for i in range(3, int(ws["AP23"].value)+4):
    print("row:",i)
    # Marks
    datalistmark = data["data" + str(i)] = []
    # Class
    info1 = ws.cell(row=i, column=1).value
    datalistinfo1 = data["Class" + str(i)] = str(info1)
    # Class no
    info2 = ws.cell(row=i, column=2).value
    datalistinfo2 = data["ClassNo" + str(i)] = str(info2)
    # Years/Date
    yearinfo = ws.cell(row=i, column=3).value
    date = data["year" + str(i)] = str(yearinfo)

    # loop
    for j in range(6, 12):
        print("col:",j)
        data1 = ws.cell(row=i, column=j).value
        # print(type(data1))
        try:
            if type(data1) == str:
                data1 = 0
                # data1 *= 0.01
            datalistmark.append(data1)
        except:
            print("An error occurs")
            pass



with open("data\data.json", "w") as f:
    json.dump(data, f)