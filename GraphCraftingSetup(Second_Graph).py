import openpyxl
import json

wb = openpyxl.load_workbook(filename="data//table_of_trial_data_for_radar_diagram.xlsx", data_only=True)
ws = wb.active

data2 = {

}

MaxRow = data2["MaxRow"] = int(ws["AP23"].value)
MaxColumn = data2["MaxColumn"] = int(ws["AP27"].value)

for i in range(3, int(ws["AP23"].value)+2):
    print("row:", i)
    datalistmark = data2["data" + str(i)] = []
    # Class
    info1 = ws.cell(row=i, column=1).value
    datalistinfo1 = data2["Class" + str(i)] = str(info1)
    # Class no
    info2 = ws.cell(row=i, column=2).value
    datalistinfo2 = data2["ClassNo" + str(i)] = str(info2)
    # Date
    dateinfo = ws.cell(row=i, column=3).value
    date = data2["date" + str(i)] = str(dateinfo)
    # Years
    yearinfo = ws.cell(row=i, column=4).value
    year = data2["year" + str(i)] = str(yearinfo)

    for j in range(15, 38):
        print("col:", j)
        data1 = ws.cell(row=i, column=j).value
        # print(type(data1))
        try:
            if type(data1) == str:
                data1 = 0
                # data1 *= 0.01
            datalistmark.append(data1)
        except:
            print("An error occurs")
            pass

with open("data\data2.json", "w") as f:
    json.dump(data2, f)